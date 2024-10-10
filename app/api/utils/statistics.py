from app.db.database import db
import calendar, datetime

import openpyxl


def get_statistics(month, year):
    statistics = {}
    total_frequency = 0
    total_enrolled = 0

    now = datetime.datetime.now()
    current_year = now.year
    current_month = now.month

    if current_year < year or (current_year == year and current_month < month):
        returned_object = {
            "daily": {},
            "monthly": {}
        }
        return returned_object, None
    # Search how many days are in a specific given month
    month_days = calendar.monthrange(datetime.datetime.now().year, month)

    # Iterate trough every day borrow to see statistics in a specific day
    daily = {}
    for day in range(1, month_days[1]+1):
        frequency = []
        returned_object = {
            "enrolled_persons": 0,
            "male_readers": 0,
            "female_readers": 0,
            "under_14": 0,
            "over_14": 0,
            "frequency": 0
        }
        if day <= now.day:
            # Check enrolled students

            enrolled_persons, error = db.get_enrolled_persons(month, year, day)
            returned_object["enrolled_persons"] = len(enrolled_persons)
            total_enrolled = total_enrolled + returned_object["enrolled_persons"]

            daily_borrows, error = db.get_daily_borrows(month, year, day)
            if daily_borrows:
                for daily_borrow in daily_borrows:
                    borrow_person = daily_borrow["person_id"]
                    if borrow_person not in frequency:
                        frequency.append(borrow_person)
                        returned_object["frequency"] += 1
                    borrow_person_info, error = db.get_person(borrow_person)

                    # Check if male of female
                    if borrow_person_info.get("gender") == "male":
                        returned_object["male_readers"] += 1
                    else:
                        returned_object["female_readers"] += 1

                    # Check if over or under 14

                    if borrow_person_info.get("year") > 8:
                        returned_object["over_14"] += 1
                    else:
                        returned_object["under_14"] += 1

                    # Total readers

                    returned_object["total_readers"] = returned_object["male_readers"] + returned_object["female_readers"]

                    daily[day] = returned_object
                total_frequency += returned_object["frequency"]
            else:
                daily[day] = returned_object
        else:
            break

    monthly_borrows, error = db.get_monthly_borrows(month, year)
    monthly = {
        "enrolled_persons": 0,
        "male_readers": 0,
        "female_readers": 0,
        "under_14": 0,
        "over_14": 0,
        "monthly_frequency": total_frequency/month_days[1]
    }
    if monthly_borrows:
        for borrow in monthly_borrows:
            borrow_person = borrow["person_id"]
            borrow_person_info, error = db.get_person(borrow_person)

            if borrow_person_info.get("gender") == "male":
                monthly["male_readers"] += 1
            else:
                monthly["female_readers"] += 1

            if borrow_person_info.get("year") > 8:
                monthly["over_14"] += 1
            else:
                monthly["under_14"] += 1

        monthly["total_readers"] = monthly["male_readers"] + monthly["female_readers"]

    monthly["enrolled_persons"] = total_enrolled

    statistics["daily"] = daily
    statistics["monthly"] = monthly

    return statistics, None


def download_statistics(month, year):
    wb = openpyxl.load_workbook("app/core/templates/template.xlsx")

    ws = wb.active

    statistics, error = get_statistics(month, year)

    print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

    template_headers = {
        "total_readers": "B",
        "workers": "C",
        "farmers": "D",
        "tehnicians": "E",
        "elevi": "F",
        "students": "G",
        "intelectuals": "H",
        "homestay": "I",
        "other": "J",
        "under_14": "K",
        "over_14": "L",
        "male_readers": "M",
        "female_readers": "N",
        "frequency": "O",
        "actions": "P",
        "participations": "Q"
    }

    daily_statistics = statistics.get("daily")

    for day in daily_statistics:
        for field in daily_statistics.get(day):
            if field != "enrolled_persons":
                cell = template_headers[field] + f"{day+4}"
                value = daily_statistics.get(day)[field]

                if int(value) == 0:
                    value = ""

                ws[cell] = value
            else:
                pass

    wb.save("app/core/files/file.xlsx")

    return None, None